import bs4, pandas as pd, openpyxl, requests, re
from openpyxl.utils.dataframe import dataframe_to_rows

results = requests.get(r'https://www.villageinsurance.com/')
print(results.status_code == requests.codes.ok)

soupy = bs4.BeautifulSoup(results.text, 'html.parser')

insurance_agent_df = pd.DataFrame({})



def html_parser(soup_object,col_name):
    object = soupy.select(soup_object)
    obj_list = []
    for i in object:
        obj_list.append(i.get_text())
    insurance_agent_df[col_name] = obj_list

#name, title and phone
html_parser('.staff-name','Name')
html_parser('.staff-title','Title')
html_parser('.staff-phone','Phone')

#getting email:
email = soupy.find_all('p',class_='staff-email')
email = [str(i) for i in email]
email_re = re.compile(r'&lt;(.*)&gt;')
email = [email_re.search(x).group(1).replace(' [at] ','@').replace(' [dot] ','.') if email_re.search(x) else '' for x in email]
print(email)
insurance_agent_df['Email'] = email

print(insurance_agent_df)

#print to excel
wb = openpyxl.Workbook()
wb.create_sheet('Village Insurance')
sheet = wb['Village Insurance']

rows = dataframe_to_rows(insurance_agent_df,index=False, header=True)

for r_idx, row in enumerate(rows, 0):
    for c_idx, value in enumerate(row, 0):
         sheet.cell(row=r_idx+1, column=c_idx+1, value=value)

wb.save(r'misc.-files\Insurance_scrape.xlsx')